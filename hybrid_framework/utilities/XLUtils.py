import openpyxl


def read_locators(path, sheet_name):

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]

    locator_dict = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if row[0]:
            locator_dict[row[0]] = [row[1], row[2]]

    workbook.close()
    return locator_dict


def read_data(path, sheet_name):

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    data_list = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_list.append(list(row))

    workbook.close()
    return data_list


def read_data_as_dicts(path, sheet_name):

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]

    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    data_list = []

    for row in rows[1:]:

        row_data = {}

        for i in range(len(headers)):
            row_data[headers[i]] = row[i]

        data_list.append(row_data)

    workbook.close()
    return data_list
